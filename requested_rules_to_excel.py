import openpyxl
from openpyxl.styles import Alignment

def convert_to_excel(data, output_file='output.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Converted Data"

    # Define column headers
    headers = ['Source', 'Destination', 'Service', 'Comments']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Start from row 2 (after headers)
    current_row = 2

    for customer, rows in data.items():
        for row in rows:
            for col, cell in enumerate(row, start=1):
                ws.cell(row=current_row, column=col, value=cell.replace('; ', '\n') if col == 4 else cell)
            current_row += 1

        # Add an empty row between data blocks
        current_row += 1

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Set text wrapping and alignment for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(customer + '_' + output_file)


if __name__ == "__main__":
    import json
    TEST_DATA = r'C:\Users\pbrehaut4\PycharmProjects\FW_Forms_pub\Sample_data\TEST_Data.json'
    with open(TEST_DATA, 'r') as file:
        data = json.load(file)

    convert_to_excel(data)