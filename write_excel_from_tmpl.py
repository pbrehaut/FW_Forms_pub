import openpyxl
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
import shutil


def write_to_excel(data, headers, field_mapping, filename="output.xlsx",
                   image_files=None, template=None):
    #  Extract the sheet to output to
    acl_sheet = headers.pop('acl_sheet', None)
    output_headers = headers.pop('output_headers', 'no')
    if output_headers.lower() == 'no':
        output_headers = False
    else:
        output_headers = True

    if template:
        # Make a copy of the template
        shutil.copy(template, filename)
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[acl_sheet] if acl_sheet else workbook.active
    else:
        # Create a new workbook if no template is provided
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = acl_sheet if acl_sheet else "ACL"

    # Set zoom level for the main sheet to 50%
    sheet.sheet_view.zoomScale = 50

    # Extract start_row and remove it from headers
    start_row = int(headers.pop('start_row', 2))

    # Add headers if output_headers is True
    if output_headers:
        for header, col_letter in headers.items():
            cell = sheet[f"{col_letter}1"]
            cell.value = header
            cell.alignment = Alignment(wrap_text=True)
    else:
        # If headers are not output, adjust start_row
        start_row -= 1

    # Write data to cells
    for row_index, row_data in enumerate(data, start=start_row):
        for header, field_index in field_mapping.items():
            if header in headers:
                col_letter = headers[header]
                cell = sheet[f"{col_letter}{row_index}"]
                cell.value = row_data[field_index]
                cell.alignment = Alignment(wrap_text=True)

    # Auto-adjust column widths based on max length of values split by '\n'
    for col_letter in headers.values():
        max_length = 0
        for cell in sheet[col_letter]:
            if cell.value:
                lines = str(cell.value).split('\n')
                for line in lines:
                    if len(line) > max_length:
                        max_length = len(line)
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[col_letter].width = adjusted_width

    # Add images to a new sheet called "Diagrams"
    if image_files:
        if "Diagrams" not in workbook.sheetnames:
            diagram_sheet = workbook.create_sheet("Diagrams")
        else:
            diagram_sheet = workbook["Diagrams"]
        for i, image_file in enumerate(image_files, start=1):
            img = Image(image_file)
            diagram_sheet.add_image(img, f'A{i * 15}')

        # Set zoom level for Diagrams sheet to 50%
        diagram_sheet.sheet_view.zoomScale = 50

    # Save the workbook
    workbook.save(filename)