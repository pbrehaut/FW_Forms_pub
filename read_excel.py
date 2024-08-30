import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import json


def read_excel_data(file_path, customer, sheet_name, start_row, source_ips, dest_ips, services, comments):
    workbook = openpyxl.load_workbook(file_path)

    if sheet_name and sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    data = {customer: []}

    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        if all(cell is None or cell == "" for cell in row):
            break

        row_data = [
            row[ord(source_ips.upper()) - 65],
            row[ord(dest_ips.upper()) - 65],
            row[ord(services.upper()) - 65],
            row[ord(comments.upper()) - 65]
        ]
        data[customer].append(row_data)

    return data

def read_excel_form():

    def process_excel():
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        customer = customer_dropdown.get()
        sheet_name = sheet_entry.get()
        start_row = int(start_row_entry.get())
        source_ips = source_ips_entry.get()
        dest_ips = dest_ips_entry.get()
        services = services_entry.get()
        comments = comments_entry.get()

        try:
            data = read_excel_data(file_path, customer, sheet_name, start_row, source_ips, dest_ips, services, comments)
            print(json.dumps(data, indent=3))  # You can process the data further or save it as needed
            messagebox.showinfo("Success", "Data processed successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    # Create and place widgets
    tk.Label(self.master, text="Customer:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
    customer_var = tk.StringVar()
    customer_dropdown = ttk.Combobox(self.master, textvariable=customer_var, values=self.customers, state="readonly")
    customer_dropdown.grid(row=0, column=1, padx=5, pady=5)
    if self.customers:
        customer_dropdown.set(self.customers[0])  # Set default value to the first customer

    tk.Label(self.master, text="Sheet Name (optional):").grid(row=1, column=0, sticky="e", padx=5, pady=5)
    sheet_entry = tk.Entry(self.master)
    sheet_entry.grid(row=1, column=1, padx=5, pady=5)

    tk.Label(self.master, text="Start Row:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
    start_row_entry = tk.Entry(self.master)
    start_row_entry.grid(row=2, column=1, padx=5, pady=5)
    start_row_entry.insert(0, "2")

    tk.Label(self.master, text="Source IPs Column:").grid(row=3, column=0, sticky="e", padx=5, pady=5)
    source_ips_entry = tk.Entry(self.master)
    source_ips_entry.grid(row=3, column=1, padx=5, pady=5)
    source_ips_entry.insert(0, "A")

    tk.Label(self.master, text="Destination IPs Column:").grid(row=4, column=0, sticky="e", padx=5, pady=5)
    dest_ips_entry = tk.Entry(self.master)
    dest_ips_entry.grid(row=4, column=1, padx=5, pady=5)
    dest_ips_entry.insert(0, "B")

    tk.Label(self.master, text="Services Column:").grid(row=5, column=0, sticky="e", padx=5, pady=5)
    services_entry = tk.Entry(self.master)
    services_entry.grid(row=5, column=1, padx=5, pady=5)
    services_entry.insert(0, "C")

    tk.Label(self.master, text="Comments Column:").grid(row=6, column=0, sticky="e", padx=5, pady=5)
    comments_entry = tk.Entry(self.master)
    comments_entry.grid(row=6, column=1, padx=5, pady=5)
    comments_entry.insert(0, "D")

    process_button = tk.Button(self.master, text="Process Excel", command=process_excel)
    process_button.grid(row=7, column=0, columnspan=2, pady=10)

