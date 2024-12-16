import openpyxl
import re
from typing import Dict, List, Tuple, Optional
from collections import Counter


def find_header_row(sheet, header_variations: Dict[str, List[str]]) -> Tuple[Optional[Dict[str, str]], int]:
    """
    Searches for header row matching specified variations.
    Returns tuple of (column_mapping, header_row) or (None, -1) if not found.
    """
    max_row = min(sheet.max_row, 20)  # Only check first 20 rows for headers
    max_column = sheet.max_column

    # Create sets of variations for each field for faster lookup
    variation_sets = {
        field: set(variations + [var.lower() for var in variations])
        for field, variations in header_variations.items()
    }

    for row in range(1, max_row + 1):
        found_columns = {}

        for col in range(1, max_column + 1):
            cell_value = str(sheet.cell(row, col).value).strip().lower()
            if not cell_value or cell_value == 'none':
                continue

            for field, variations in variation_sets.items():
                if cell_value in variations:
                    found_columns[field] = openpyxl.utils.get_column_letter(col)
                    break

        # If we found all required headers
        if len(found_columns) == len(header_variations):
            return found_columns, row

    return None, -1


def analyze_excel_workbook(file_path: str) -> Dict[str, Dict[str, str]]:
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    result = {}

    # Define header variations to look for
    header_variations = {
        'source_ips': ['source', 'sources', 'src', 'source ip', 'source ips'],
        'dest_ips': ['destination', 'destinations', 'dst', 'dest', 'destination ip', 'destination ips'],
        'services': ['port', 'ports', 'service', 'services', 'protocol', 'protocols'],
        'comments': ['comment', 'comments', 'description', 'descriptions', 'notes']
    }

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # First try to find headers
        column_mapping, header_row = find_header_row(sheet, header_variations)

        if column_mapping:
            # If headers found, add the start row (one after header)
            column_mapping['start_row'] = str(header_row + 1)
            result[sheet_name] = column_mapping
        else:
            # Fall back to automatic detection
            sheet_data = analyze_sheet(sheet)
            if sheet_data:
                result[sheet_name] = sheet_data

    return result


def analyze_sheet(sheet) -> Dict[str, str]:
    max_row = sheet.max_row
    max_column = sheet.max_column

    ip_pattern = re.compile(r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}(?:/[0-9]{1,3}|_[0-9]{1,3})?\b')
    services = ['HTTP', 'HTTPS', 'DNS', 'FTP', 'SMTP', 'POP3', 'IMAP', 'SSH', 'TELNET', 'RDP']
    protocols = ['TCP', 'UDP', 'ICMP', 'IP']

    potential_rules = []

    for row in range(1, max_row + 1):
        rule = analyze_row(sheet, row, max_column, ip_pattern, services, protocols)
        if rule:
            potential_rules.append(rule)

    if not potential_rules:
        return {}

    # Find the most common column for each field
    field_columns = {
        'source_ips': Counter(rule['source_ips'] for rule in potential_rules).most_common(1)[0][0],
        'dest_ips': Counter(rule['dest_ips'] for rule in potential_rules).most_common(1)[0][0],
        'services': Counter(rule['services'] for rule in potential_rules).most_common(1)[0][0],
        'comments': Counter(rule['comments'] for rule in potential_rules).most_common(1)[0][0]
    }

    # Find the start row
    start_row = min(rule['row'] for rule in potential_rules)

    field_columns['start_row'] = str(start_row)
    return field_columns


def analyze_row(sheet, row: int, max_column: int, ip_pattern, services: List[str], protocols: List[str]) -> Dict[
    str, str]:
    source_ips = None
    dest_ips = None
    ports_or_services = None
    comments = None

    for col in range(1, max_column + 1):
        cell_value = str(sheet.cell(row, col).value).strip()
        if not cell_value or cell_value == 'None':
            continue

        if not source_ips and ip_pattern.search(cell_value):
            source_ips = openpyxl.utils.get_column_letter(col)
        elif not dest_ips and ip_pattern.search(cell_value):
            dest_ips = openpyxl.utils.get_column_letter(col)
        elif not ports_or_services and (any(service.lower() in cell_value.lower() for service in services) or
                                        any(protocol.lower() in cell_value.lower() for protocol in protocols)):
            ports_or_services = openpyxl.utils.get_column_letter(col)
        elif not comments and len(cell_value.split()) > 0:
            comments = openpyxl.utils.get_column_letter(col)

    if all([source_ips, dest_ips, ports_or_services, comments]):
        return {
            'row': row,
            'source_ips': source_ips,
            'dest_ips': dest_ips,
            'services': ports_or_services,
            'comments': comments
        }
    return None


# Usage example
if __name__ == "__main__":
    file_path = ""
    result = analyze_excel_workbook(file_path)
    print(result)