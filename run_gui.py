import os
import tkinter as tk
from tkinter import ttk, messagebox, Text, filedialog, simpledialog
import json
import generate_xls_diagrams
from configmanager import ConfigManager
import configparser
import openpyxl
import find_rules_excel

# Global variable for config file
CONFIG_FILE = 'config.ini'


class NetworkInfoGUI:
    def __init__(self, master):
        self.master = master
        self.master.title("Network Information")
        self.master.geometry("600x600")

        self.stored_file_path = tk.StringVar()
        self.stored_diagram_dir_path = tk.StringVar()

        self.topology = ConfigManager(CONFIG_FILE)
        self.customers = self.topology.get_customers()
        self.results = {}
        self.selected_customer = tk.StringVar()
        self.file_path = tk.StringVar()
        self.customer_combo = None

        # Initial Form
        self.create_initial_form()

    def read_excel_data(self, file_path, customer, sheet_name, start_row, source_ips, dest_ips, services, comments):
        workbook = openpyxl.load_workbook(file_path)

        if sheet_name and sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        data = {customer: []}

        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            if all(cell is None or cell == "" for cell in row):
                break

            row_data = [
                row[ord(source_ips.upper()) - 65],
                row[ord(dest_ips.upper()) - 65],
                row[ord(services.upper()) - 65],
                row[ord(comments.upper()) - 65]
            ]
            data[customer].append(row_data)

        return data

    def read_excel_auto_form(self):
        excel_window = tk.Toplevel(self.master)
        excel_window.title("Read Excel Data - Auto-detect")
        excel_window.geometry("300x200")

        tk.Label(excel_window, text="Customer:").pack(pady=5)
        customer_var = tk.StringVar()
        customer_dropdown = ttk.Combobox(excel_window, textvariable=customer_var, values=self.customers,
                                         state="readonly")
        customer_dropdown.pack(pady=5)
        if self.customers:
            customer_dropdown.set(self.customers[0])

        def process_excel_auto():
            config_mgr = ConfigManager(CONFIG_FILE)
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            if not file_path:
                return

            sheets_rule_spec = find_rules_excel.analyze_excel_workbook(file_path)
            if not sheets_rule_spec:
                messagebox.showerror("Error", "No valid sheets found.")
                return
            for sheet_name, rule_spec in sheets_rule_spec.items():
                customer = customer_var.get()
                start_row = rule_spec['start_row']
                source_ips = rule_spec['source_ips']
                dest_ips = rule_spec['dest_ips']
                services = rule_spec['services']
                comments = rule_spec['comments']

                try:
                    self.results = self.read_excel_data(file_path, customer, sheet_name, int(start_row), source_ips,
                                                        dest_ips, services, comments)
                    self.process_results()
                except Exception as e:
                    messagebox.showerror("Error", f"An error occurred: {str(e)}")

            excel_window.destroy()
            self.create_initial_form()

        process_button = tk.Button(excel_window, text="Process Excel", command=process_excel_auto)
        process_button.pack(pady=10)

    def read_excel_form(self):
        excel_window = tk.Toplevel(self.master)
        excel_window.title("Read Excel Data")
        excel_window.geometry("300x500")

        tk.Label(excel_window, text="Customer:").pack(pady=5)
        customer_var = tk.StringVar()
        customer_dropdown = ttk.Combobox(excel_window, textvariable=customer_var, values=self.customers, state="readonly")
        customer_dropdown.pack(pady=5)
        if self.customers:
            customer_dropdown.set(self.customers[0])

        tk.Label(excel_window, text="Sheet Name (optional):").pack(pady=5)
        sheet_entry = tk.Entry(excel_window)
        sheet_entry.pack(pady=5)

        tk.Label(excel_window, text="Start Row:").pack(pady=5)
        start_row_entry = tk.Entry(excel_window)
        start_row_entry.pack(pady=5)
        start_row_entry.insert(0, "2")

        tk.Label(excel_window, text="Source IPs Column:").pack(pady=5)
        source_ips_entry = tk.Entry(excel_window)
        source_ips_entry.pack(pady=5)
        source_ips_entry.insert(0, "A")

        tk.Label(excel_window, text="Destination IPs Column:").pack(pady=5)
        dest_ips_entry = tk.Entry(excel_window)
        dest_ips_entry.pack(pady=5)
        dest_ips_entry.insert(0, "B")

        tk.Label(excel_window, text="Services Column:").pack(pady=5)
        services_entry = tk.Entry(excel_window)
        services_entry.pack(pady=5)
        services_entry.insert(0, "C")

        tk.Label(excel_window, text="Comments Column:").pack(pady=5)
        comments_entry = tk.Entry(excel_window)
        comments_entry.pack(pady=5)
        comments_entry.insert(0, "D")

        def process_excel():
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            if not file_path:
                return

            customer = customer_var.get()
            sheet_name = sheet_entry.get()
            start_row = int(start_row_entry.get())
            source_ips = source_ips_entry.get()
            dest_ips = dest_ips_entry.get()
            services = services_entry.get()
            comments = comments_entry.get()

            try:
                self.results = self.read_excel_data(file_path, customer, sheet_name, start_row, source_ips, dest_ips, services, comments)
                messagebox.showinfo("Success", "Data processed successfully!")
                self.load_excel_to_manual_form()
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred: {str(e)}")

        process_button = tk.Button(excel_window, text="Process Excel", command=process_excel)
        process_button.pack(pady=10)

    def edit_config_file(self):
        # Read the current contents of the config file
        with open(CONFIG_FILE, 'r') as file:
            content = file.read()

        # Create a new top-level window
        edit_window = tk.Toplevel(self.master)
        edit_window.title("Edit Config File")
        edit_window.geometry("700x600")

        # Create a Text widget with a scrollbar
        text_frame = tk.Frame(edit_window)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        text_widget = Text(text_frame, wrap=tk.NONE)
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=text_widget.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        text_widget.config(yscrollcommand=scrollbar.set)

        # Insert the content into the Text widget
        text_widget.insert(tk.END, content)

        # Function to save the changes
        def save_changes():
            new_content = text_widget.get("1.0", tk.END)
            with open(CONFIG_FILE, 'w') as file:
                file.write(new_content)
            messagebox.showinfo("Success", "Config file updated successfully")
            edit_window.destroy()

        # Add a save button
        save_button = tk.Button(edit_window, text="Save Changes", command=save_changes)
        save_button.pack(pady=10)

    def refresh_customers(self):
        self.topology = ConfigManager(CONFIG_FILE)
        self.customers = self.topology.get_customers()
        self.customer_combo['values'] = list(self.customers)
        if self.selected_customer.get() not in self.customers:
            self.selected_customer.set('')

    def go_back_to_initial_form(self):
        # Clear the manual input form
        self.manual_input_frame.destroy()

        # Recreate the initial form
        self.create_initial_form()

    def browse_directory(self):
        dir_path = filedialog.askdirectory()
        if dir_path:
            self.stored_diagram_dir_path.set(dir_path)

    def render_diagrams(self):
        if not self.stored_diagram_dir_path.get():
            messagebox.showwarning("Warning", "Please select a directory.")
            return

        from diagram_renderer import render_diagrams_in_directory
        render_diagrams_in_directory(self.stored_diagram_dir_path.get())
        messagebox.showinfo("Rendering Complete", "Diagrams have been rendered successfully.")

    def create_initial_form(self):
        self.master.geometry("600x400")
        self.master.grid_columnconfigure(0, weight=1)
        self.master.grid_columnconfigure(1, weight=1)

        # File Selection Section
        file_frame = tk.LabelFrame(self.master, text="Load from File", font=("Arial", 12, "bold"))
        file_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        self.file_entry = tk.Entry(file_frame, textvariable=self.stored_file_path, width=30)
        self.file_entry.grid(row=0, column=0, padx=5, pady=5)

        self.browse_button = tk.Button(file_frame, text="Browse", command=self.browse_file)
        self.browse_button.grid(row=0, column=1, padx=5, pady=5)

        self.load_button = tk.Button(file_frame, text="Load JSON File", command=self.load_file)
        self.load_button.grid(row=1, column=0, columnspan=2, pady=5)

        # Excel Load Section
        excel_frame = tk.LabelFrame(self.master, text="Load from Excel", font=("Arial", 12, "bold"))
        excel_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

        self.load_excel_button = tk.Button(excel_frame, text="Load From Excel", command=self.read_excel_form)
        self.load_excel_button.grid(row=0, column=0, pady=5)

        self.load_auto_excel_button = tk.Button(excel_frame, text="Auto Load From Excel",
                                                command=self.read_excel_auto_form)
        self.load_auto_excel_button.grid(row=1, column=0, pady=5)

        # Customer Selection Section
        customer_frame = tk.LabelFrame(self.master, text="Manual Input", font=("Arial", 12, "bold"))
        customer_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

        self.select_customer_label = tk.Label(customer_frame, text="Select Customer:")
        self.select_customer_label.grid(row=0, column=0, pady=5)

        self.customer_combo = ttk.Combobox(customer_frame, textvariable=self.selected_customer,
                                           values=list(self.customers))
        self.customer_combo.grid(row=1, column=0, pady=5)

        self.start_manual_button = tk.Button(customer_frame, text="Start Manual Input", command=self.start_manual_input)
        self.start_manual_button.grid(row=2, column=0, pady=5)

        # Diagram Rendering Section
        diagram_frame = tk.LabelFrame(self.master, text="Render Diagrams", font=("Arial", 12, "bold"))
        diagram_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")

        self.diagram_dir_entry = tk.Entry(diagram_frame, textvariable=self.stored_diagram_dir_path, width=30)
        self.diagram_dir_entry.grid(row=0, column=0, padx=5, pady=5)

        self.browse_dir_button = tk.Button(diagram_frame, text="Browse", command=self.browse_directory)
        self.browse_dir_button.grid(row=0, column=1, padx=5, pady=5)

        self.render_button = tk.Button(diagram_frame, text="Render Diagrams", command=self.render_diagrams)
        self.render_button.grid(row=1, column=0, columnspan=2, pady=5)

        # Additional Buttons
        button_frame = tk.Frame(self.master)
        button_frame.grid(row=2, column=0, columnspan=2, pady=10)

        self.modify_config_button = tk.Button(button_frame, text="Add/Modify\nCustomers/Topologies",
                                              command=self.modify_config)
        self.modify_config_button.pack(side=tk.LEFT, padx=5)

        self.edit_config_button = tk.Button(button_frame, text=f"Edit Config File:\n{CONFIG_FILE}",
                                            command=self.edit_config_file)
        self.edit_config_button.pack(side=tk.LEFT, padx=5)

    def create_manual_input_form(self):
        #  Reset self.results to an empty dictionary
        # self.results = {}

        self.master.geometry("700x600")

        self.manual_input_frame = ttk.Frame(self.master)
        self.manual_input_frame.pack(padx=10, pady=10, fill="both", expand=True)

        # Input Form
        input_form = ttk.LabelFrame(self.manual_input_frame, text="Input Form")
        input_form.pack(padx=10, pady=10, fill="both", expand=True)

        self.input_fields = {}
        field_heights = {"Source IPs": 5, "Destination IPs": 5, "Services": 2, "Comments": 3}

        first_field = None

        for field, height in field_heights.items():
            frame = ttk.Frame(input_form)
            frame.pack(fill="x", padx=5, pady=5)
            ttk.Label(frame, text=f"{field}:").pack(anchor="w", padx=5)
            text = Text(frame, height=height, width=40)
            text.pack(fill="x", padx=5)
            self.input_fields[field] = text

            # Make Text widgets tabbable
            text.configure(takefocus=1)

            # Store the first field
            if first_field is None:
                first_field = text

        # Buttons
        button_frame = ttk.Frame(input_form)
        button_frame.pack(fill="x", padx=5, pady=5)
        ttk.Button(button_frame, text="Add Entry", command=self.add_entry).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Update Entry", command=self.update_entry).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Delete Entry", command=self.delete_entry).pack(side=tk.LEFT, padx=5)

        # Add a "Back" button
        back_button = ttk.Button(button_frame, text="Back", command=self.go_back_to_initial_form)
        back_button.pack(side=tk.LEFT, padx=5)

        # Add an "Options" button
        options_button = ttk.Button(button_frame, text="Options",
                                    command=lambda: self.edit_excel_form(self.selected_customer.get() + ".EXCEL",
                                                                    CONFIG_FILE))
        options_button.pack(side=tk.LEFT, padx=5)

        # Add an "template select" button
        template_button = ttk.Button(button_frame, text="Template",
                                    command=lambda: self.edit_excel_form(self.selected_customer.get() + ".FILES",
                                                                    CONFIG_FILE))
        template_button.pack(side=tk.LEFT, padx=5)

        # Submit Button (now in button_frame)
        self.submit_button = ttk.Button(button_frame, text="Submit", command=self.submit_results)
        self.submit_button.pack(side=tk.RIGHT, padx=5)

        # Assuming this is part of a class, and self.manual_input_frame already exists
        # Create a frame to hold the Treeview and scrollbar
        tree_frame = tk.Frame(self.manual_input_frame)
        tree_frame.pack(padx=10, pady=10, fill="both", expand=True)

        # Create the Treeview widget
        self.tree = ttk.Treeview(tree_frame,
                                 columns=("Source IPs", "Destination IPs", "Services", "Comments"), show="headings")
        self.tree.heading("Source IPs", text="Source IPs")
        self.tree.column("Source IPs", width=0, stretch=tk.NO)
        self.tree.heading("Destination IPs", text="Destination IPs")
        self.tree.column("Destination IPs", width=0, stretch=tk.NO)
        self.tree.heading("Services", text="Services")
        self.tree.column("Services", width=0, stretch=tk.NO)
        self.tree.heading("Comments", text="Comments")
        self.tree.column("Comments", width=200)

        # Create vertical scrollbar
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)

        # Grid layout for Treeview and scrollbar
        self.tree.grid(column=0, row=0, sticky='nsew')
        vsb.grid(column=1, row=0, sticky='ns')

        # Configure the tree_frame grid
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        # Bind double-click event
        self.tree.bind("<Double-1>", self.on_tree_double_click)

        # Set tab order
        self.bind_tab_navigation()

        # Load data into Treeview if provided
        if self.results:
            data = self.results.get(self.selected_customer.get(), [])
            for entry in data:
                self.tree.insert("", "end", values=entry)

        # Set focus to the first field
        if first_field:
            self.master.after(100, lambda: first_field.focus_set())

    def bind_tab_navigation(self):
        def focus_next_widget(event):
            event.widget.tk_focusNext().focus()
            return "break"

        def focus_previous_widget(event):
            event.widget.tk_focusPrev().focus()
            return "break"

        for field in self.input_fields.values():
            field.bind("<Tab>", focus_next_widget)
            field.bind("<Shift-Tab>", focus_previous_widget)

        self.tree.bind("<Tab>", focus_next_widget)
        self.tree.bind("<Shift-Tab>", focus_previous_widget)

        for button in [self.submit_button]:  # Add other buttons if needed
            button.bind("<Tab>", focus_next_widget)
            button.bind("<Shift-Tab>", focus_previous_widget)

    def browse_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("JSON Files", "*.json"), ("All Files", "*.*")])
        if file_path:
            self.stored_file_path.set(file_path)

    def load_file(self):
        if not self.stored_file_path.get():
            messagebox.showwarning("Warning", "Please select a file.")
            return
        try:
            with open(self.stored_file_path.get(), 'r') as file:
                self.results = json.load(file)

            # Extract the customer key from the JSON data
            customer_key = list(self.results.keys())[0]
            self.selected_customer.set(customer_key)

            # Remove initial form
            for widget in self.master.winfo_children():
                widget.destroy()

            # Create and show manual input form
            self.create_manual_input_form()

            messagebox.showinfo("File Loaded", f"File {self.stored_file_path.get()} loaded successfully.")
        except json.JSONDecodeError:
            messagebox.showerror("Error", "Invalid JSON file. Please select a valid JSON file.")
        except IOError:
            messagebox.showerror("Error",
                                 "Could not read the file. Please check if the file exists and you have permission to read it.")

    def load_excel_to_manual_form(self):
        # Extract the customer key from the JSON data
        customer_key = list(self.results.keys())[0]
        self.selected_customer.set(customer_key)

        # Remove initial form
        for widget in self.master.winfo_children():
            widget.destroy()

        # Create and show manual input form
        self.create_manual_input_form()

    def process_results(self):
        config_mgr = ConfigManager(CONFIG_FILE)
        user_msg = generate_xls_diagrams.generate_output(self.results, config_mgr)

        if not user_msg:
            user_msg = "No IPs unmatched to a topology."

        header = f'''Requested rules processed and output to:\n{config_mgr.get_output_directory(self.selected_customer.get())}\n\n'''

        user_msg = header + user_msg

        # Create a new window for the custom input form
        result_window = tk.Toplevel(self.master)
        result_window.title("Output")
        result_window.geometry("600x400")

        # Create a frame to hold the text box
        text_frame = tk.Frame(result_window)
        text_frame.pack(padx=10, pady=10, fill="both", expand=True)

        # Create the text box
        result_text = Text(text_frame, width=80, height=15, font=("Arial", 12))
        result_text.pack(fill="both", expand=True)

        # Insert the user_msg into the text box
        result_text.insert("1.0", user_msg)

        # Create a button to close the window
        close_button = tk.Button(result_window, text="Close", command=result_window.destroy)
        close_button.pack(pady=10)

        # Set the text box as read-only
        result_text.configure(state="disabled")

        # Bind the Enter key to close the window
        result_window.bind('<Return>', lambda event: result_window.destroy())

        # Set focus to the result window so it can receive keyboard events
        result_window.focus_set()

    def submit_results(self):
        self.results[self.selected_customer.get()] = [self.tree.item(item)["values"] for item in
                                                      self.tree.get_children()]
        self.process_results()

    def start_manual_input(self):
        if not self.selected_customer.get():
            messagebox.showwarning("Warning", "Please select a customer.")
            return

        # Clear previous data
        self.results = {}
        self.file_path.set("")
        self.stored_file_path.set("")
        self.stored_diagram_dir_path.set("")

        # Remove initial form
        for widget in self.master.winfo_children():
            widget.destroy()

        # Create and show manual input form
        self.create_manual_input_form()
        self.results[self.selected_customer.get()] = []

    def add_entry(self):
        entry = {field: self.input_fields[field].get("1.0", tk.END).strip() for field in self.input_fields}
        self.tree.insert("", "end", values=(entry["Source IPs"], entry["Destination IPs"], entry["Services"], entry["Comments"].replace('\n', '; ')))
        self.clear_input_fields()

    def update_entry(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select an entry to update.")
            return

        entry = {field: self.input_fields[field].get("1.0", tk.END).strip() for field in self.input_fields}
        self.tree.item(selected_item, values=(entry["Source IPs"], entry["Destination IPs"], entry["Services"], entry["Comments"].replace('\n', '; ')))
        self.clear_input_fields()

    def delete_entry(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select an entry to delete.")
            return
        self.tree.delete(selected_item)

    def on_tree_double_click(self, event):
        item = self.tree.selection()[0]
        values = self.tree.item(item, "values")
        for field, value in zip(self.input_fields.keys(), values):
            self.input_fields[field].delete("1.0", tk.END)
            if field == "Comments":
                value = value.replace('; ', '\n')
            self.input_fields[field].insert("1.0", value)

    def clear_input_fields(self):
        for text in self.input_fields.values():
            text.delete("1.0", tk.END)

    def edit_excel_form(self, section, config_file):
        config = configparser.ConfigParser()
        config.read(config_file)

        # Refresh config manager with updated file
        self.topology = ConfigManager(CONFIG_FILE)

        # Create a new window for editing the section
        edit_window = tk.Toplevel(self.master)
        edit_window.title(f"Edit {section} Section")

        # Define a dictionary to hold the tkinter variables
        var_dict = {}

        # Add widgets for each option in the section
        for i, (option, value) in enumerate(config[section].items()):
            label = tk.Label(edit_window, text=option)
            label.grid(row=i, column=0, padx=10, pady=5, sticky="e")

            if option == "template_filename":
                # Populate the drop-down with filenames from the template directory
                files = self.topology.get_files_config(self.selected_customer.get())
                files = files.get('template_directory', None)
                files = os.listdir(files)
                files.insert(0,"None")
                var_dict[option] = tk.StringVar(value=value)
                option_menu = ttk.Combobox(edit_window, textvariable=var_dict[option], values=files, width=40)
                option_menu.grid(row=i, column=1, padx=10, pady=5)
            elif option in ["group_gateways", "detailed_diagrams", "include_flow_count", "output_headers"]:
                var_dict[option] = tk.StringVar(value=value)
                option_menu = ttk.Combobox(edit_window, textvariable=var_dict[option], values=["yes", "no"])
                option_menu.grid(row=i, column=1, padx=10, pady=5)
            else:
                var_dict[option] = tk.StringVar(value=value)
                entry = tk.Entry(edit_window, textvariable=var_dict[option])
                entry.grid(row=i, column=1, padx=10, pady=5)

        # Save button
        def save_changes():
            for option, var in var_dict.items():
                config[section][option] = var.get()
            with open(config_file, 'w') as configfile:
                config.write(configfile)
            edit_window.destroy()

        save_button = tk.Button(edit_window, text="Save", command=save_changes)
        save_button.grid(row=len(config[section]), column=0, columnspan=2, pady=10)

    def modify_config(self):
        config = configparser.ConfigParser()
        if os.path.exists(CONFIG_FILE):
            config.read(CONFIG_FILE)

        customer = simpledialog.askstring("Input", "Enter customer name:")
        if not customer:
            return

        if customer not in config:
            # New customer: proceed with full configuration
            config[customer] = {}
            files_section = f"{customer}.FILES"
            config[files_section] = {}
            config[files_section]['template_filename'] = 'None'

            for dir_type in ['topology_directory', 'template_directory', 'output_directory']:
                dir_path = filedialog.askdirectory(title=f"Select {dir_type.replace('_', ' ')} for {customer}")
                if dir_path:
                    config[files_section][dir_type] = dir_path
                else:
                    config[files_section][dir_type] = 'None'

            # Add EXCEL section with predefined options
            excel_section = f"{customer}.EXCEL"
            config[excel_section] = {
                'group_gateways': 'yes',
                'detailed_diagrams': 'no',
                'diagram_node_comments': 'no',
                'diagram_max_ips': '3',
                'include_flow_count': 'no',
                'output_headers': 'yes',
                'acl_sheet': 'ACL',
                'start_row': '2',
                'source_ips': 'A',
                'destination_ips': 'B',
                'services': 'C',
                'comments': 'D',
                'gateway': 'E',
                'rule_id': 'F',
                'paths': 'G'
            }

        # Always prompt for topology, whether it's a new or existing customer
        topologies = []
        while True:
            topology = simpledialog.askstring("Input", f"Enter topology name for {customer} (or cancel to finish):")
            if not topology:
                break

            topology_data = {'name': topology, 'files': {}}
            for file_type in ['subnets', 'routes', 'topology']:
                file_path = filedialog.askopenfilename(title=f"Select {file_type} file for {topology}")
                if file_path:
                    topology_data['files'][file_type] = os.path.basename(file_path)
                else:
                    topology_data['files'][file_type] = 'None'
            topologies.append(topology_data)

        if topologies:
            if f"{customer}.TOPOLOGIES" not in config:
                config[f"{customer}.TOPOLOGIES"] = {}
            for topology in topologies:
                topology_section = f"{customer}.TOPOLOGIES.{topology['name']}"
                config[topology_section] = topology['files']

        with open(CONFIG_FILE, 'w') as configfile:
            config.write(configfile)

        messagebox.showinfo("Success", "Config file updated successfully!")

        # Refresh the customer list and combo box
        self.refresh_customers()


def run_gui():
    root = tk.Tk()
    gui = NetworkInfoGUI(root)
    root.mainloop()
    return gui.results


if __name__ == "__main__":
    run_gui()
